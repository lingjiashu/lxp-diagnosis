"""
Excel 解析器：解析三种逆变器数据文件
1. 历史运行数据 (per-day sheets, 114 columns)
2. 事件记录 (Event Record sheet)
3. 操作设置记录 (Remote Set Record sheet)
"""

import pandas as pd
from datetime import datetime
from typing import Optional
import re


def _extract_serial_from_xls(filepath: str) -> str:
    """用 xlrd/openpyxl 直接读取原始序列号（避免科学计数法问题）"""
    # Try xlrd for .xls files
    if filepath.lower().endswith('.xls') and not filepath.lower().endswith('.xlsx'):
        try:
            import xlrd
            wb = xlrd.open_workbook(filepath)
            sheet = wb.sheet_by_index(0)
            header_row = 0
            sn_col = None
            for col in range(sheet.ncols):
                if str(sheet.cell_value(header_row, col)).strip() == 'Serial number':
                    sn_col = col
                    break
            if sn_col is not None:
                for row in range(1, sheet.nrows):
                    cell = sheet.cell(row, sn_col)
                    if cell.ctype == xlrd.XL_CELL_TEXT:
                        return cell.value.strip()
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        raw = str(cell.value)
                        return raw
        except ImportError:
            pass
    
    # Try openpyxl for .xlsx files
    if filepath.lower().endswith('.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheet = wb.worksheets[0]
            sn_col = None
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value and str(cell.value).strip() == 'Serial number':
                    sn_col = col_idx
                    break
            if sn_col is not None:
                for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 10),
                                           min_col=sn_col, max_col=sn_col):
                    cell = row[0]
                    if cell.value is not None:
                        val = str(cell.value).strip()
                        return val
            wb.close()
        except ImportError:
            pass
    
    return None


def parse_historical_data(filepath: str) -> dict:
    """
    解析逆变器历史运行数据 Excel。
    每个 sheet 是一天的数据，列名包含 114 个字段。
    
    返回:
    {
        "serial_number": str,
        "sheets": ["2025-04-24", ...],
        "total_rows": int,
        "time_range": {"start": datetime, "end": datetime},
        "rows": [list of dicts],  # 所有数据合并
        "columns": [str],         # 列名列表
    }
    """
    # First try to extract serial number from raw xls
    serial_number = _extract_serial_from_xls(filepath)
    
    xl = pd.ExcelFile(filepath)
    all_rows = []
    start_time = None
    end_time = None
    columns = None
    
    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name)
        if df.empty:
            continue
        
        if columns is None:
            columns = list(df.columns)
        
        # Fallback serial number extraction
        if serial_number is None and 'Serial number' in df.columns:
            sn_val = df['Serial number'].iloc[0]
            if pd.notna(sn_val):
                serial_number = str(sn_val)
        
        # Parse time
        if 'Time' in df.columns:
            for t in df['Time']:
                if pd.notna(t):
                    try:
                        dt = pd.to_datetime(str(t))
                        if start_time is None or dt < start_time:
                            start_time = dt
                        if end_time is None or dt > end_time:
                            end_time = dt
                    except Exception:
                        pass
        
        # Convert rows to dict
        for _, row in df.iterrows():
            row_dict = row.to_dict()
            # Parse time
            if 'Time' in row_dict and pd.notna(row_dict['Time']):
                try:
                    row_dict['_parsed_time'] = pd.to_datetime(str(row_dict['Time']))
                except Exception:
                    row_dict['_parsed_time'] = None
            all_rows.append(row_dict)
    
    return {
        "serial_number": serial_number,
        "sheets": xl.sheet_names,
        "total_rows": len(all_rows),
        "time_range": {"start": start_time, "end": end_time},
        "rows": all_rows,
        "columns": columns or [],
    }


def parse_event_records(filepath: str) -> dict:
    """
    解析事件记录 Excel。
    
    列: Station, Serial number, Event Type, Event, Start Time, Time Recovered
    
    返回:
    {
        "serial_number": str,
        "events": [
            {
                "station": str,
                "type": "Fault" | "Notice",
                "event": str,
                "start_time": datetime,
                "recovered_time": datetime or None,
            }
        ],
        "fault_events": [...],   # 仅 Fault 类型
        "notice_events": [...],  # 仅 Notice 类型
    }
    """
    xl = pd.ExcelFile(filepath)
    sheet_name = xl.sheet_names[0]  # Usually "Event Record"
    df = xl.parse(sheet_name)
    
    serial_number = None
    events = []
    
    for _, row in df.iterrows():
        if serial_number is None and pd.notna(row.get('Serial number')):
            sn_val = row['Serial number']
            serial_number = str(int(sn_val)) if pd.notna(sn_val) else None
        
        start_t = None
        recovered_t = None
        
        if pd.notna(row.get('Start Time')):
            try:
                start_t = pd.to_datetime(str(row['Start Time']))
            except Exception:
                pass
        
        if pd.notna(row.get('Time Recovered')):
            try:
                recovered_t = pd.to_datetime(str(row['Time Recovered']))
            except Exception:
                pass
        
        events.append({
            "station": str(row.get('Station', '')),
            "type": str(row.get('Event Type', '')),
            "event": str(row.get('Event', '')),
            "start_time": start_t,
            "recovered_time": recovered_t,
        })
    
    fault_events = [e for e in events if e['type'] == 'Fault']
    notice_events = [e for e in events if e['type'] == 'Notice']
    
    return {
        "serial_number": serial_number,
        "events": events,
        "fault_events": fault_events,
        "notice_events": notice_events,
        "total_events": len(events),
    }


def parse_set_records(filepath: str) -> dict:
    """
    解析操作设置记录 Excel。
    
    列: Time, Username, Station, Serial number, Dongle, Client Type, 
        Set Type, Set Result, Parameter Name, Parameter Value
    
    返回:
    {
        "serial_number": str,
        "records": [
            {
                "time": datetime,
                "username": str,
                "station": str,
                "client_type": str,
                "set_type": str,
                "result": str,  # "Success" | "Failed"
                "parameter": str,
                "value": str,
            }
        ]
    }
    """
    xl = pd.ExcelFile(filepath)
    sheet_name = xl.sheet_names[0]
    df = xl.parse(sheet_name)
    
    serial_number = None
    records = []
    
    for _, row in df.iterrows():
        if serial_number is None and pd.notna(row.get('Serial number')):
            sn_val = row['Serial number']
            serial_number = str(int(sn_val)) if pd.notna(sn_val) else None
        
        t = None
        if pd.notna(row.get('Time')):
            try:
                t = pd.to_datetime(str(row['Time']))
            except Exception:
                pass
        
        records.append({
            "time": t,
            "username": str(row.get('Username', '')),
            "station": str(row.get('Station', '')),
            "client_type": str(row.get('Client Type', '')),
            "set_type": str(row.get('Set Type', '')),
            "result": str(row.get('Set Result', '')),
            "parameter": str(row.get('Parameter Name', '')),
            "value": str(row.get('Parameter Value', '')),
        })
    
    return {
        "serial_number": serial_number,
        "records": records,
        "total_records": len(records),
    }


def safe_float(val) -> Optional[float]:
    """安全转换为 float"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if pd.isna(val):
            return None
        return float(val)
    try:
        s = str(val).strip()
        if s == '' or s.lower() == 'nan':
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def safe_int(val) -> Optional[int]:
    """安全转换为 int，支持 0x 前缀"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if pd.isna(val):
            return None
        return int(val)
    try:
        s = str(val).strip()
        if s == '' or s.lower() == 'nan':
            return None
        if s.startswith('0x') or s.startswith('0X'):
            return int(s, 16)
        return int(float(s))
    except (ValueError, TypeError):
        return None
